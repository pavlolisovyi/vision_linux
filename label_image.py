import os, sys
import datetime
import openpyxl

import tensorflow as tf
def main():
    os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'

    # change this as you see fit
    image_path = "cropped.jpg"

    # Read in the image_data
    image_data = tf.gfile.FastGFile(image_path, 'rb').read()

    # Loads label file, strips off carriage return
    label_lines = [line.rstrip() for line 
                    in tf.gfile.GFile("retrained_labels.txt")]

    # Unpersists graph from file
    with tf.gfile.FastGFile("retrained_graph.pb", 'rb') as f:
        graph_def = tf.GraphDef()
        graph_def.ParseFromString(f.read())
        tf.import_graph_def(graph_def, name='')

    with tf.Session() as sess:
        # Feed the image_data as input to the graph and get first prediction
        softmax_tensor = sess.graph.get_tensor_by_name('final_result:0')
    
        predictions = sess.run(softmax_tensor, \
                {'DecodeJpeg/contents:0': image_data})
    
        # Sort to show labels of first prediction in order of confidence
        top_k = predictions[0].argsort()[-len(predictions[0]):][::-1]
       

        #Handling the excel file
        wb = openpyxl.load_workbook('sample.xlsx')
        # grab the active worksheet
        ws = wb.active

        num_loops = 0
        input_row = ws.max_row + 1
        ws['G{}'.format(input_row)] = datetime.datetime.now()
        #outputing the result of label_image.py
        for node_id in top_k:
            human_string = label_lines[node_id]
            score = predictions[0][node_id]
            print('%s (score = %.5f)' % (human_string, score))
            if num_loops == 0:
                ws['A{}'.format(input_row)] = score
                ws['B{}'.format(input_row)] = human_string
            elif num_loops == 1:
                ws['C{}'.format(input_row)] = score
                ws['D{}'.format(input_row)] = human_string
            elif num_loops == 2:
                ws['E{}'.format(input_row)] = score
                ws['F{}'.format(input_row)] = human_string
            else:
                pass
            num_loops = num_loops + 1
        #Saving the excel file
        wb.save("sample.xlsx")